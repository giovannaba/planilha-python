import openpyxl

# criar uma planilha(book)
book = openpyxl.Workbook()

#  como visualizar paginas existentes
print(book.sheetnames)

# como criar uma pagina
book.create_sheet('Frutas')

# como selecionar uma pagina
frutas_page = book['Frutas']

# colunas
frutas_page.append(['Fruta', 'Qtd', 'Valor'])
frutas_page.append(['Banan', '5','R$3.90'])
frutas_page.append(['Banan', '5','R$3.90'])
frutas_page.append(['Banan', '5','R$3.90'])
frutas_page.append(['Banan', '5','R$3.90'])
frutas_page.append(['Bananaaa', '5','R$3.90'])

# salvar planilha
book.save('Planilha de compras.xlsx')